"""Build '[UploadedName]_ExperimentalResults.xlsx'.

One sheet per experiment. Each sheet:
  - header block (metric, algorithm, reference, gains, run/wait time) at top-left
  - achieved performance (IAE/ITAE/ISE/ITSE computed from the measured run)
    directly beneath it
  - the two graphs (motor speed vs time, motor voltage vs time) pinned to the
    upper-right, next to that block
  - the full data table below
"""

import io
import os

import numpy as np
import matplotlib
matplotlib.use("Agg")            # headless: render to PNG, never open a window
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage

import plant_model


DATA_HEADERS = ["time [s]", "rpm", "rpm_ref", "motor_input",
                "PWM [%]", "voltage [V]", "rpm_raw"]
DATA_KEYS    = ["time", "rpm", "rpm_ref", "motor_input", "pwm", "voltage", "rpm_raw"]

DATA_HEADER_ROW = 19             # row where the data-table header goes (cols A-G)
# 2x2 graph grid to the right of the header/data (cols I & Q), measured on top,
# expected (model) below.
IMG_ANCHOR_SPEED       = "I1"    # measured motor speed
IMG_ANCHOR_VOLTAGE     = "Q1"    # measured motor voltage
IMG_ANCHOR_EXP_SPEED   = "I18"   # expected motor speed (plant model)
IMG_ANCHOR_EXP_VOLTAGE = "Q18"   # expected motor voltage (plant model)

TITLE_FONT  = Font(bold=True, size=13)
LABEL_FONT  = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(bold=True, color="FFFFFF")


def _trapz(y, x):
    fn = getattr(np, "trapezoid", None) or np.trapz
    return float(fn(y, x))


def compute_metrics(time_s, rpm, reference):
    """IAE / ITAE / ISE / ITSE from the measured error e(t) = ref - rpm(t)."""
    t = np.asarray(time_s, dtype=float)
    y = np.asarray(rpm, dtype=float)
    e = float(reference) - y
    ae = np.abs(e)
    se = e * e
    return {
        "IAE":  _trapz(ae, t),
        "ITAE": _trapz(t * ae, t),
        "ISE":  _trapz(se, t),
        "ITSE": _trapz(t * se, t),
    }


def _make_chart_png(x, y, title, ylabel, ref=None, ref_label=None):
    fig, ax = plt.subplots(figsize=(4.7, 3.0), dpi=100)
    ax.plot(x, y, linewidth=1.3)
    if ref is not None:
        ax.axhline(ref, linestyle="--", linewidth=1.0, color="tab:red",
                   label=ref_label or f"ref = {ref}")
        ax.legend(loc="lower right", fontsize=8)
    ax.set_title(title, fontsize=10)
    ax.set_xlabel("time [s]", fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.tick_params(labelsize=8)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


def _safe_sheet_name(name, used):
    # Excel: <=31 chars, no []:*?/\
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    name = name[:31] or "Sheet"
    base, i = name, 1
    while name.lower() in used:
        suffix = f"_{i}"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _write_sheet(ws, result, model=None):
    data = result["data"]
    ref  = float(result["reference"])
    metrics = compute_metrics(data["time"], data["rpm"], ref)

    # ---- header block (col A label, col B value) ----
    ws["A1"] = f"EXPERIMENT: {result['metric']} / {result['algorithm']}"
    ws["A1"].font = TITLE_FONT

    info = [
        ("Optimization Metric", result["metric"]),
        ("Algorithm",           result["algorithm"]),
        ("Reference [RPM]",      ref),
        ("Kp",                   result["kp"]),
        ("Ki",                   result["ki"]),
        ("Kd",                   result["kd"]),
        ("Run time [s]",         result.get("run_time")),
        ("Wait time [s]",        result.get("wait_time")),
    ]
    r = 3
    for label, value in info:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1

    # ---- achieved performance ----
    r += 1  # blank spacer (row 12)
    c = ws.cell(row=r, column=1, value="Achieved performance (from measured data)")
    c.font = LABEL_FONT
    c.fill = SECTION_FILL
    ws.cell(row=r, column=2).fill = SECTION_FILL
    r += 1
    for key in ("IAE", "ITAE", "ISE", "ITSE"):
        ws.cell(row=r, column=1, value=key).font = LABEL_FONT
        ws.cell(row=r, column=2, value=round(metrics[key], 4))
        r += 1

    # ---- measured graphs (top row of the grid) ----
    speed_png = _make_chart_png(
        data["time"], data["rpm"], "Motor Speed vs Time (measured)", "speed [rpm]",
        ref=ref, ref_label=f"reference = {ref:g} rpm")
    volt_png = _make_chart_png(
        data["time"], data["voltage"], "Motor Voltage vs Time (measured)", "voltage [V]")
    ws.add_image(XLImage(speed_png), IMG_ANCHOR_SPEED)
    ws.add_image(XLImage(volt_png), IMG_ANCHOR_VOLTAGE)

    # ---- expected graphs from the plant model (bottom row of the grid) ----
    tvals = [float(x) for x in data["time"] if x is not None]
    tfinal = max(tvals) if tvals else 5.0
    t_model = np.arange(0.0, tfinal + 0.01, 0.01)

    # optional edited plant transfer function from the GUI
    if model:
        plant_num = [float(model["num"])]
        plant_den = [1.0, float(model["a1"]), float(model["a0"])]
        Tf   = float(model["Tf"])
        vmax = float(model["vmax"])
    else:
        plant_num = plant_den = Tf = None
        vmax = plant_model.VMAX

    try:
        rpm_exp, volt_exp = plant_model.expected_response(
            float(result["kp"]), float(result["ki"]), float(result["kd"]),
            ref, t_model, plant_num=plant_num, plant_den=plant_den, Tf=Tf)
        exp_speed_png = _make_chart_png(
            t_model, rpm_exp, "Expected Motor Speed vs Time (model)", "speed [rpm]",
            ref=ref, ref_label=f"reference = {ref:g} rpm")
        exp_volt_png = _make_chart_png(
            t_model, volt_exp, "Expected Motor Voltage vs Time (model)", "voltage [V]",
            ref=vmax, ref_label=f"Vmax = {vmax:g} V")
        ws.add_image(XLImage(exp_speed_png), IMG_ANCHOR_EXP_SPEED)
        ws.add_image(XLImage(exp_volt_png), IMG_ANCHOR_EXP_VOLTAGE)
    except Exception as e:                      # never fail the whole workbook
        ws["I18"] = f"Expected-response model failed: {e}"

    # ---- data table ----
    hr = DATA_HEADER_ROW
    for j, htext in enumerate(DATA_HEADERS, start=1):
        cell = ws.cell(row=hr, column=j, value=htext)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    n = len(data["time"])
    for i in range(n):
        row = hr + 1 + i
        for j, key in enumerate(DATA_KEYS, start=1):
            ws.cell(row=row, column=j, value=data[key][i])

    # column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for col in "CDEFG":
        ws.column_dimensions[col].width = 12


def build_results_workbook(results, out_path, model=None):
    """results: list of 'result' dicts from comm_client. Writes out_path.

    model (optional): edited plant transfer-function parameters from the GUI
    ({num, a1, a0, Tf, vmax}) for the expected-response graphs. When None, the
    defaults from compare_all_methods.m are used.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    used = set()
    ordered = sorted(results, key=lambda r: r.get("index", 0))
    for res in ordered:
        name = _safe_sheet_name(f"{res['metric']}_{res['algorithm']}", used)
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, res, model)

    wb.save(out_path)
    return out_path


def default_output_name(input_xlsx_path):
    base = os.path.splitext(os.path.basename(input_xlsx_path))[0]
    folder = os.path.dirname(os.path.abspath(input_xlsx_path))
    return os.path.join(folder, f"{base}_ExperimentalResults.xlsx")
