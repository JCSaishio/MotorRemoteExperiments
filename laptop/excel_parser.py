"""Read the 'Summary' sheet of an uploaded Excel file into a list of experiments.

Expected columns (header row, case-insensitive): Metric, Algorithm, Kp, Ki, Kd
(extra columns such as BestCost / Iterations / Runtime are ignored).
"""

import openpyxl


REQUIRED = ["metric", "algorithm", "kp", "ki", "kd"]


def parse_experiments(xlsx_path, sheet_name="Summary"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"'{sheet_name}' sheet not found. Sheets in file: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"'{sheet_name}' sheet is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(name):
        for i, h in enumerate(header):
            if h == name:
                return i
        return None

    idx = {name: col(name) for name in REQUIRED}
    missing = [n for n in REQUIRED if idx[n] is None]
    if missing:
        raise ValueError(
            f"Missing column(s) {missing} in '{sheet_name}'. Header found: {rows[0]}")

    experiments = []
    n = 0
    for r in rows[1:]:
        metric = r[idx["metric"]]
        algo   = r[idx["algorithm"]]
        kp     = r[idx["kp"]]
        # skip blank / separator rows
        if metric is None and algo is None:
            continue
        if kp is None:
            continue
        n += 1
        experiments.append({
            "index":     n,
            "metric":    str(metric).strip(),
            "algorithm": str(algo).strip(),
            "kp":        float(r[idx["kp"]]),
            "ki":        float(r[idx["ki"]]),
            "kd":        float(r[idx["kd"]]),
        })

    if not experiments:
        raise ValueError(f"No experiment rows found in '{sheet_name}'.")
    return experiments


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "../Example.xlsx"
    for e in parse_experiments(path):
        print(e)
